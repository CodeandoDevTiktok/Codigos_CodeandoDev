import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

libro = openpyxl.Workbook()
hoja = libro.active

#Elimina lineas excel
hoja.sheet_view.showGridLines = False

# Data

data = [['Nombre','Apellido','Edad'],
        ['Juan','Perez',32],
        ['Maria','Gonzales',28],
        ['Pedro','Garcia',45]]

#Estilo fuente
estilo_fuente1 = Font(name='Arial',size=12,bold=True)
estilo_fuente2 = Font(name='Arial',size=10,italic=True)

#Borde
borde = Border(left=Side(border_style='thin',color='000000'),
               right=Side(border_style='thin',color='000000'),
               top=Side(border_style='thin',color='000000'),
               bottom=Side(border_style='thin',color='000000'))

#Relleno
relleno1 = PatternFill(fill_type='solid',fgColor='00FFFFFF')
relleno2 = PatternFill(fill_type='solid',fgColor='0099CCFF')

#Alineacion
alinacion = Alignment(horizontal='center',vertical='center')

for fila in data:
    hoja.append(fila)
    for col in range(1,len(fila) + 1):
        celda = hoja.cell(row=hoja.max_row,column=col)
        celda.font = estilo_fuente1 if hoja.max_row == 1 else estilo_fuente2
        celda.alignment = alinacion
        celda.border = borde
        celda.fill = relleno1 if hoja.max_row % 2 == 0 else relleno2

#Ajustar ancho de columnas
for columna_celda in hoja.columns:
    longitud = max(len(str(cell.value)) for cell in columna_celda)
    hoja.column_dimensions[columna_celda[0].column_letter].width = longitud*1.80

libro.save('datos.xlsx')

