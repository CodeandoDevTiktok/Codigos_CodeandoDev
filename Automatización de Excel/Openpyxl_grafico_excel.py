import openpyxl
import matplotlib.pyplot as plt
from openpyxl.styles import NamedStyle, Alignment

# Crear un archivo de Excel nuevo
libro = openpyxl.Workbook()

# Crear una hoja de Excel en el libro y dar formato
hoja = libro.active
hoja.title = 'Ventas'
hoja.sheet_view.showGridLines = False
hoja.column_dimensions['A'].width = 15
hoja.column_dimensions['B'].width = 20
hoja.column_dimensions['C'].width = 15
hoja.column_dimensions['D'].width = 15
estilo_cabecera = NamedStyle(name='cabecera')
estilo_cabecera.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
estilo_cabecera.font = estilo_cabecera.font.__class__(bold=True)
hoja['A1'].style = estilo_cabecera
hoja['B1'].style = estilo_cabecera
hoja['C1'].style = estilo_cabecera
hoja['D1'].style = estilo_cabecera

# Crear una lista de ventas con sus respectivos campos
ventas = [
    {'Fecha': '2022-01-01', 'Producto': 'Laptop', 'Cantidad': 3, 'Precio': 900},
    {'Fecha': '2022-01-02', 'Producto': 'Smartphone', 'Cantidad': 5, 'Precio': 600},
    {'Fecha': '2022-01-03', 'Producto': 'Tablet', 'Cantidad': 2, 'Precio': 400},
    {'Fecha': '2022-01-04', 'Producto': 'Smartwatch', 'Cantidad': 1, 'Precio': 200},
    {'Fecha': '2022-01-05', 'Producto': 'TV', 'Cantidad': 1, 'Precio': 1200}
]

# Escribir los datos de ventas en la hoja de Excel
hoja.cell(row=1, column=1).value = 'Fecha'
hoja.cell(row=1, column=2).value = 'Producto'
hoja.cell(row=1, column=3).value = 'Cantidad'
hoja.cell(row=1, column=4).value = 'Precio'
fila = 2
for venta in ventas:
    hoja.cell(row=fila, column=1).value = venta['Fecha']
    hoja.cell(row=fila, column=2).value = venta['Producto']
    hoja.cell(row=fila, column=3).value = venta['Cantidad']
    hoja.cell(row=fila, column=4).value = venta['Precio']
    fila += 1

# Crear un gráfico de barras utilizando la librería matplotlib
datos = hoja['C2:C6']
productos = hoja['B2:B6']
x = [p[0].value for p in productos]
y = [d[0].value for d in datos]

plt.bar(x, y)
plt.xlabel('Productos')
plt.ylabel('Cantidad')
plt.title('Ventas por producto')

# Crear una nueva hoja de Excel para el grafico
hoja_grafico = libro.create_sheet('Gráfico')

#Agregar el gráfico a la nueva hoja de Excel
grafico = openpyxl.chart.BarChart()
grafico.title = 'Ventas por producto'
grafico.x_axis.title = 'Productos'
grafico.y_axis.title = 'Cantidad'
grafico.add_data(openpyxl.chart.Reference(hoja, min_col=3, min_row=2, max_col=3, max_row=6))
grafico.set_categories(openpyxl.chart.Reference(hoja, min_col=2, min_row=2, max_col=2, max_row=6))
hoja_grafico.add_chart(grafico, 'A1')

#Guardar el archivo de Excel
libro.save('ventas.xlsx')