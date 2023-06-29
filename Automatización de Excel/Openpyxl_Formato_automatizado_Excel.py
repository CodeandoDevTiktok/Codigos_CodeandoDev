from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment,Border,Side

#Cargar Excel

libro = load_workbook('data_excel_2022.xlsx')

#Hoja

hoja = libro['FSI 2006-2022']

#Formato :

#Sin lineas
hoja.sheet_view.showGridLines = False

#Borde

borde = Border(left=Side(border_style="thin",color="000000"),
               right=Side(border_style="thin",color="000000"),
               top=Side(border_style="thin",color="000000"),
               bottom=Side(border_style="thin",color="000000"))

#Formato a celdas:

for fila in hoja.iter_rows():
    for celda in fila:
        if celda.value is not None:
            celda.font = Font(bold=True,size=8)
            celda.border = borde
            celda.alignment = Alignment(horizontal="center", vertical="center")
            
#Guardar cambios

libro.save("data_excel_2022_con_formato.xlsx")            